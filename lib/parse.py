from openpyxl import load_workbook
import pandas as pd
import xlrd

file = ''
workbook = None

def set_file_name(file_name):
    file = file_name

def get_file_name():
    return file

def load_work(file_name):
    set_file_name(file_name)
    global file
    file = file_name
    global workbook
    workbook = load_workbook(filename=file_name)

def get_sheet_by_index(sheet_by_id):
    return workbook.sheet_by_index(sheet_by_id)

def get_sheet_names():
    return workbook.sheetnames

def get_the_data_from_sheet(sheet_name, forceDouble, eliminateString):
    excel_data = ''
    global var_name
    var_name = ''
    sheet_ranges = workbook[sheet_name]
    var_fun = ''

    if eliminateString.isChecked() == True and forceDouble.isChecked() == True:
        var_name = 'std::vector<std::vector<double>> {p1}_array'.format(p1=sheet_name) + '\n{'
    else:
        var_name = 'std::array<std::array<std::any,{p1}>,{p2}> {p3}_array'.format(p1=sheet_ranges.max_column, p2=sheet_ranges.max_row, p3=sheet_name) + '\n{'
        var_fun = 'std::array<std::any,{p1}>'.format(p1=sheet_ranges.max_column)

    if forceDouble.isChecked() == True:
        wb = load_workbook(filename=file, data_only=True)
        sheet_ranges = wb[sheet_name]

    excel = ''
    for col in sheet_ranges.iter_rows():
        excel = '\n{p1}'.format(p1=var_fun) + '{ '
        for cell in col:
            if type(cell.value) == int or type(cell.value) == float:
                excel = excel  + str(cell.value) + ','
            elif type(cell.value) == str:
                dat = cell.value
                dat = dat.replace('//','\/\/')
                dat = dat.replace('"','\\"')
                dat = dat.replace('\n','')
                if eliminateString.isChecked() == True:
                    if '=' in dat and forceDouble.isChecked() == False:
                        excel = excel + "\"" + dat + '\",'
                    else:
                        excel = excel + str(0.0) + ","
                else:
                    excel = excel + "\"" + dat + "\","
            else:
                excel = excel + str(0.0) + ','

        excel = excel[:-1]
        excel = excel + ' },'
        excel_data = excel_data + excel

    excel_data = excel_data[:-1]
    excel_data = var_name + excel_data + '\n};\n'
    return excel_data
